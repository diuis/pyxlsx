import xlsxwriter
import os
from openpyxl import load_workbook

workDirectory = './xlsx-data'

if not os.path.exists(workDirectory):
    os.makedirs(workDirectory)

# creating test files

filesToCreate = range(1, 256)
for fileToCreate in filesToCreate:
    if not os.path.isfile(workDirectory + '/file_' + str(fileToCreate) + '.xlsx'):
        print('writing file number ' + str(fileToCreate))
        newXslx = xlsxwriter.Workbook(
            workDirectory + '/file_' + str(fileToCreate) + '.xlsx', {'constant_memory': False})
        newXslxSheet = newXslx.add_worksheet()
        rows = range(1, 15000)
        for row in rows:
            columns = range(1, 6)
            for column in columns:
                newXslxSheet.write(
                    row, column, 'la nebbia agli irti colli piovigginando sale')
        newXslx.close()


# creating multisheets file from test files

convertedXslx = xlsxwriter.Workbook(
    workDirectory + '/multisheets.xlsx', {'constant_memory': True})

sheets = range(1, 256)
for sheetItem in sheets:

    xslxToConvert = load_workbook(
        filename=workDirectory + '/file_' + str(sheetItem) + '.xlsx', read_only=True)
    xslxToConvertSheet = xslxToConvert.active

    worksheet = convertedXslx.add_worksheet()
    for rowIndex, row in enumerate(xslxToConvertSheet.values):
        print('row is ' + str(row))
        for colIndex, column in enumerate(row):
            print('column is ' + str(column))
            worksheet.write(rowIndex, colIndex, str(column))

convertedXslx.close()
